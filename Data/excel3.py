"""
使用openpyxl实现以下需求
使用excel 写入一组数据，姓名，身高，体重
计算是否为健康体重，如果是健康体重，则在旁边备注健康，并将姓名打印出来
健康体重计算公式：（身高cm-70）×60%
(可以做一部分优化)
"""
from openpyxl import Workbook,load_workbook


class PracticeExcel:
    def create_data(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "INFO"
        ws1["A1"]="姓名"
        ws1["B1"]="身高"
        ws1["C1"]="体重"
        name=["Lily","Lucy","Lilei","Jerry","Jack","Sara"]
        height=[165,163,178,175,182,162]
        weight=[53,52,80,63,90,50]
        for i in range(len(name)):
            ws1.cell(row=i+2, column=1).value = name[i]
            ws1.cell(row=i+2, column=2).value = height[i]
            ws1.cell(row=i+2, column=3).value = weight[i]
        wb.save(filename="data.xlsx")
    def read_data(self):
        lw=load_workbook(filename="data.xlsx")
        sheet=lw["INFO"]
        for i in range(7):
            print(sheet.cell(row=i + 1, column=1).value)
    def health_data(self):
        lw=load_workbook(filename="data.xlsx")
        sheet=lw["INFO"]
        sheet["C1"]="备注"
        for i in range(6):
            height = sheet.cell(row=i + 2, column=2).value
            weight = sheet.cell(row=i + 2, column=3).value
            helth_weight = (height - 70) * 0.6
            #print("健康体重是",helth_weight)
            if weight==helth_weight:
                print("健康体重是",weight)
                sheet.cell(row=i + 2, column=4, value= "健康体重")
                # sheet.cell(row=i + 2, column=4).value="健康体重"
        lw.save(filename="data.xlsx")

pe = PracticeExcel()
#pe.create_data()
pe.health_data()
