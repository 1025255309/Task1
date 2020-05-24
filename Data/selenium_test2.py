import time

from selenium import webdriver
from openpyxl import load_workbook

class SeleniumPython:

    def __init__(self):
        ld = load_workbook("./selenium.xlsx")
        self.ws = ld.worksheets[0]

    def go_baidu(self):
        """
        1. 打开百度
        2. 输入搜索词
        3. 点击搜索
        将元素定位和输入的数据使用excel管理
        :retur        n:
        """
        # 实例化driver
        driver = webdriver.Chrome()
        driver.get("https://www.baidu.com/")
        time.sleep(2)
        driver.find_element_by_id(self.ws["A2"].value).send_keys(self.ws["B2"].value)
        driver.find_element_by_id(self.ws["A3"].value).click()
        time.sleep(15)

if __name__ == '__main__':
    sp = SeleniumPython()
    sp.go_baidu()